import glob
import openpyxl
from openpyxl import load_workbook

# tested using openpyxl 3.0.3 on python 3.7.5, ubuntu 18.04

# TODO this script should be merged in with data/processed/pre-process-fit.py

headercols = None
rowswritten = 0

def formatcell(item):
	"Appropriate conversions for CSV output"
	if item.value is None:
		return ""
	return str(item.internal_value)

with open("fit.csv", "wt") as outfp:
	for infname in sorted(glob.glob("installation_report_*_part_*.xlsx")):
		print(infname)
		wb = load_workbook(filename=infname, read_only=True)

		gotheader = False
		for whichrow, row in enumerate(wb.active.iter_rows()):
			at_repeated_header = False
			try:
				rowstrs = [formatcell(item) for item in row]
			except AttributeError as err:
				print(err)
				print("    skipping row %i (merged cell?)" % whichrow)
				continue
			for item in rowstrs:
				assert("," not in item)
			# Here we're detecting hitting the header in this XLS:
			if rowstrs[0] and (rowstrs[0].startswith("Extension")):
				gotheader = True
				#print(headercols)
				if headercols:
					assert(headercols==rowstrs)
					at_repeated_header = True
				else:
					headercols = rowstrs
			if gotheader and not at_repeated_header:
				#if whichrow < 20:
				#	print(rowstrs[0])
				outfp.write(",".join(rowstrs) + "\n")
				rowswritten += 1
		del wb

print("Written %i rows" % rowswritten)
assert(rowswritten > 10000)
assert(gotheader)

